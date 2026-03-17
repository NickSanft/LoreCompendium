import hashlib
import logging
import os
import time
import queue
import threading
import json
from typing import List, NotRequired, Optional
import msoffcrypto  # noqa
import openpyxl  # noqa
import unstructured  # noqa
import docx  # noqa

# Concurrency imports
import concurrent.futures

from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from langchain_ollama import ChatOllama, OllamaEmbeddings
from pydantic import BaseModel, Field
from typing_extensions import TypedDict

from langchain_community.document_loaders import (
    UnstructuredWordDocumentLoader,
    PyPDFLoader,
    UnstructuredExcelLoader,
    CSVLoader,
    TextLoader
)
from langchain_community.vectorstores.utils import filter_complex_metadata
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_chroma import Chroma
from langchain_core.documents import Document
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser
from langgraph.graph import END, StateGraph, START

from lore_utils import CHROMA_COLLECTION_NAME, CHROMA_DB_PATH, THINKING_OLLAMA_MODEL, FAST_OLLAMA_MODEL, \
    EMBEDDING_MODEL, SUPPORTED_EXTENSIONS, DOC_FOLDER

logger = logging.getLogger(__name__)

GLOBAL_VECTORSTORE: Optional[Chroma] = None


def _format_location(meta: dict) -> str:
    """Convert chunk metadata into a human-readable location string."""
    if "line_start" in meta:
        ls, le = meta["line_start"], meta.get("line_end", meta["line_start"])
        return f"Line {ls}" if ls == le else f"Lines {ls}–{le}"
    if meta.get("sheet") and "row" in meta:
        return f"Sheet '{meta['sheet']}' · Row {meta['row']}"
    if "row" in meta:
        return f"Row {meta['row']}"
    if "paragraph_index" in meta:
        return f"Paragraph {meta['paragraph_index']}"
    if "page" in meta:
        return f"Page {meta['page'] + 1}"
    if "start_index" in meta:
        return f"Char {meta['start_index']}"
    return "Unknown"


def _citation_location(meta: dict) -> str:
    """Location string for citations: prefer page number, then line, then other."""
    if "page" in meta:
        return f"Page {meta['page'] + 1}"
    if "line_start" in meta:
        ls, le = meta["line_start"], meta.get("line_end", meta["line_start"])
        return f"Line {ls}" if ls == le else f"Lines {ls}–{le}"
    if meta.get("sheet") and "row" in meta:
        return f"Sheet '{meta['sheet']}' · Row {meta['row']}"
    if "row" in meta:
        return f"Row {meta['row']}"
    if "paragraph_index" in meta:
        return f"Paragraph {meta['paragraph_index']}"
    if "start_index" in meta:
        return f"Char {meta['start_index']}"
    return ""


def _enrich_line_numbers(splits: list) -> list:
    """Add line_start and line_end to splits from .txt/.md files.

    The loader stores a JSON-encoded newline-offset list in
    ``_line_offsets_json`` metadata. This function reads it, computes line
    numbers via bisect, writes line_start / line_end, then removes the
    temporary key so it is not stored in ChromaDB.
    """
    import bisect
    for doc in splits:
        raw = doc.metadata.pop("_line_offsets_json", None)
        if raw is None:
            continue
        try:
            offsets = json.loads(raw)
        except Exception:
            continue
        start = doc.metadata.get("start_index", 0)
        end = start + len(doc.page_content)
        doc.metadata["line_start"] = bisect.bisect_right(offsets, start)
        doc.metadata["line_end"]   = bisect.bisect_right(offsets, max(end - 1, start))
    return splits
INGESTION_QUEUE = queue.Queue()
MANIFEST_LOCK = threading.Lock()
INDEXED_FILES_PATH = os.path.join(CHROMA_DB_PATH, "indexed_files.txt")

_EMBEDDINGS: Optional[OllamaEmbeddings] = None
_OBSERVER: Optional[Observer] = None


def _get_embeddings() -> OllamaEmbeddings:
    """Returns the shared OllamaEmbeddings instance, creating it once on first call."""
    global _EMBEDDINGS
    if _EMBEDDINGS is None:
        _EMBEDDINGS = OllamaEmbeddings(model=EMBEDDING_MODEL)
    return _EMBEDDINGS


def load_pdf(file_path: str) -> List[Document]:
    loader = PyPDFLoader(file_path)
    docs = loader.load()
    return docs


def _load_xlsx(file_path: str) -> List[Document]:
    """Load an Excel workbook row-by-row using openpyxl, preserving sheet/row metadata."""
    import openpyxl
    docs = []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                row_num = row[0].row
                text = "\t".join(
                    str(cell.value) for cell in row if cell.value is not None
                ).strip()
                if text:
                    docs.append(Document(
                        page_content=text,
                        metadata={"source": file_path, "sheet": sheet_name, "row": row_num},
                    ))
        wb.close()
    except Exception as e:
        logger.error(f"Error loading xlsx {file_path}: {e}")
    return docs


def load_document_by_extension(file_path: str) -> List[Document]:
    """
    Selects the appropriate loader based on file extension.
    Top-level function for pickling support in multiprocessing.
    """
    # Safety check for watcher race conditions
    if not os.path.exists(file_path):
        return []

    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == '.pdf':
            return load_pdf(file_path)
        elif ext == '.docx':
            loader = UnstructuredWordDocumentLoader(file_path, mode="elements")
            docs = loader.load()
            for i, doc in enumerate(docs):
                # unstructured stores page numbers as 'page_number' (1-indexed);
                # remap to 'page' (0-indexed) so _citation_location picks it up
                if "page_number" in doc.metadata and "page" not in doc.metadata:
                    pn = doc.metadata.pop("page_number")
                    if pn is not None:
                        doc.metadata["page"] = int(pn) - 1
                doc.metadata["paragraph_index"] = i + 1
            return docs
        elif ext == '.xlsx':
            return _load_xlsx(file_path)
        elif ext == '.csv':
            loader = CSVLoader(file_path)
            return loader.load()
        elif ext in ['.txt', '.md']:
            loader = TextLoader(file_path, encoding='utf-8', autodetect_encoding=True)
            docs = loader.load()
            # Pre-compute newline offsets for post-split line number enrichment
            if docs:
                raw_text = docs[0].page_content
                offsets = [0]
                for line in raw_text.splitlines(keepends=True):
                    offsets.append(offsets[-1] + len(line))
                offsets_json = json.dumps(offsets)
                for doc in docs:
                    doc.metadata["_line_offsets_json"] = offsets_json
            return docs
        return []
    except Exception as e:
        logger.error(f"Error loading {file_path}: {e}")
        return []


class DocumentEventHandler(FileSystemEventHandler):
    """
    Watchdog Event Handler.
    Pushes events to a thread-safe queue to be processed by the worker.
    """

    def on_created(self, event):
        if not event.is_directory and event.src_path.lower().endswith(SUPPORTED_EXTENSIONS):
            logger.info(f"[WATCHER] File Created: {event.src_path}")
            INGESTION_QUEUE.put(("add", event.src_path))

    def on_modified(self, event):
        if not event.is_directory and event.src_path.lower().endswith(SUPPORTED_EXTENSIONS):
            logger.info(f"[WATCHER] File Modified: {event.src_path}")
            INGESTION_QUEUE.put(("update", event.src_path))

    def on_deleted(self, event):
        if not event.is_directory and event.src_path.lower().endswith(SUPPORTED_EXTENSIONS):
            logger.info(f"[WATCHER] File Deleted: {event.src_path}")
            INGESTION_QUEUE.put(("delete", event.src_path))

    def on_moved(self, event):
        if not event.is_directory:
            if event.src_path.lower().endswith(SUPPORTED_EXTENSIONS):
                logger.info(f"[WATCHER] File Moved (Delete old): {event.src_path}")
                INGESTION_QUEUE.put(("delete", event.src_path))
            if event.dest_path.lower().endswith(SUPPORTED_EXTENSIONS):
                logger.info(f"[WATCHER] File Moved (Add new): {event.dest_path}")
                INGESTION_QUEUE.put(("add", event.dest_path))


def _get_file_signature(file_path: str) -> dict:
    hasher = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            hasher.update(chunk)
    return {
        "mtime": os.path.getmtime(file_path),
        "size": os.path.getsize(file_path),
        "hash": hasher.hexdigest(),
    }


def get_duplicate_source(file_path: str) -> Optional[str]:
    """
    Returns the basename of an already-indexed file whose content is identical
    to file_path, or None if no duplicate exists.
    """
    if not os.path.exists(file_path):
        return None
    sig = _get_file_signature(file_path)
    new_hash = sig.get("hash")
    manifest = _load_index_manifest()
    for indexed_path, stored_sig in manifest.items():
        if indexed_path == file_path:
            continue
        if stored_sig.get("hash") == new_hash:
            return os.path.basename(indexed_path)
    return None


def _load_index_manifest() -> dict:
    if not os.path.exists(INDEXED_FILES_PATH):
        return {}
    try:
        with open(INDEXED_FILES_PATH, "r", encoding="utf-8") as f:
            raw = f.read().strip()
        if not raw:
            return {}
        if raw.startswith("{"):
            return json.loads(raw)
        # Backward compatibility with legacy newline list
        return {line.strip(): {} for line in raw.splitlines() if line.strip()}
    except Exception as e:
        logger.error(f"Error reading index manifest: {e}")
        return {}


def _write_index_manifest(manifest: dict) -> None:
    try:
        with open(INDEXED_FILES_PATH, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)
    except Exception as e:
        logger.error(f"Error writing index manifest: {e}")


def _update_manifest(action: str, file_path: str, tags: Optional[list] = None) -> None:
    with MANIFEST_LOCK:
        manifest = _load_index_manifest()
        if action == "delete":
            manifest.pop(file_path, None)
            _write_index_manifest(manifest)
            return
        if os.path.exists(file_path):
            existing_tags = manifest.get(file_path, {}).get("tags", [])
            sig = _get_file_signature(file_path)
            sig["tags"] = tags if tags is not None else existing_tags
            manifest[file_path] = sig
            _write_index_manifest(manifest)


def get_tags(file_path: str) -> list[str]:
    """Return the tag list for a file from the manifest, or [] if not found."""
    return _load_index_manifest().get(file_path, {}).get("tags", [])


def set_tags(file_path: str, tags: list[str]) -> None:
    """Replace the tag list for a file and re-queue it for ingestion so
    ChromaDB chunk metadata is refreshed with the new tags."""
    cleaned = [t.lower().strip()[:50] for t in tags if t.strip()][:20]
    with MANIFEST_LOCK:
        manifest = _load_index_manifest()
        if file_path not in manifest:
            return
        manifest[file_path]["tags"] = cleaned
        _write_index_manifest(manifest)
    INGESTION_QUEUE.put(("update", file_path))


def ingestion_worker():
    """
    Background daemon that consumes the queue.
    Handles 'add', 'update', and 'delete' operations on the global VectorStore.
    """
    global GLOBAL_VECTORSTORE

    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=1000,
        chunk_overlap=200,
        add_start_index=True
    )

    while True:
        try:
            # Block until an item is available; None is the shutdown sentinel
            item = INGESTION_QUEUE.get()
            if item is None:
                INGESTION_QUEUE.task_done()
                break
            action, file_path = item

            # Wait briefly to let file writes settle (debounce)
            time.sleep(1.0)

            if GLOBAL_VECTORSTORE is None:
                INGESTION_QUEUE.task_done()
                continue

            base_name = os.path.basename(file_path)

            logger.info(f"SYNC WORKER: Processing {action} for {base_name}")

            if action == "delete":
                try:
                    GLOBAL_VECTORSTORE.delete(where={"source": file_path})
                    _update_manifest("delete", file_path)
                    logger.info(f"Removed chunks for {base_name}")
                except Exception as e:
                    logger.error(f"Error deleting {file_path}: {e}")

            elif action in ["add", "update"]:
                # For update, we delete first to avoid duplicates
                if action == "update":
                    try:
                        GLOBAL_VECTORSTORE.delete(where={"source": file_path})
                    except Exception:
                        pass  # Might not exist yet

                # Load and Ingest
                try:
                    docs = load_document_by_extension(file_path)
                    if docs:
                        splits = text_splitter.split_documents(docs)
                        splits = _enrich_line_numbers(splits)
                        # Read current tags from manifest and stamp onto each split
                        manifest_snap = _load_index_manifest()
                        for split in splits:
                            src = split.metadata.get("source", "")
                            t = manifest_snap.get(src, {}).get("tags", [])
                            if t:
                                split.metadata["tags"] = ",".join(t)
                        splits = filter_complex_metadata(splits)
                        if splits:
                            GLOBAL_VECTORSTORE.add_documents(splits)
                            _update_manifest("add", file_path)
                            logger.info(f"Added {len(splits)} chunks for {os.path.basename(file_path)}")
                except Exception as e:
                    logger.error(f"Error ingesting {file_path}: {e}")

            INGESTION_QUEUE.task_done()

        except Exception as e:
            logger.error(f"Worker error: {e}")


# --- PART 1: OPTIMIZED INGESTION ENGINE ---

def initialize_vectorstore():
    """
    Ingests documents using Multiprocessing and returns the VectorStore.
    Also starts the background Watchdog observer.
    """
    global GLOBAL_VECTORSTORE, _OBSERVER

    embeddings = _get_embeddings()

    # Ensure directories exist
    if not os.path.exists(DOC_FOLDER):
        os.makedirs(DOC_FOLDER)
    if not os.path.exists(CHROMA_DB_PATH):
        os.makedirs(CHROMA_DB_PATH)

    # 1. Connect to VectorStore
    vectorstore = Chroma(
        persist_directory=CHROMA_DB_PATH,
        embedding_function=embeddings,
        collection_name=CHROMA_COLLECTION_NAME
    )
    GLOBAL_VECTORSTORE = vectorstore

    # 2. Identify New Files
    manifest = _load_index_manifest()
    indexed_files = set(manifest.keys())
    current_files = {}
    for root, dirs, files in os.walk(DOC_FOLDER):
        for file in files:
            if file.startswith("~$"): continue
            if file.lower().endswith(SUPPORTED_EXTENSIONS):
                path = os.path.join(root, file)
                current_files[path] = _get_file_signature(path)

    current_paths = set(current_files.keys())
    deleted_files = list(indexed_files - current_paths)
    candidate_files = list(current_paths)

    new_files = []
    updated_files = []
    for file_path in candidate_files:
        if file_path not in manifest:
            new_files.append(file_path)
            continue
        if manifest.get(file_path) != current_files[file_path]:
            updated_files.append(file_path)

    # 3. Parallel Loading & Ingestion
    if deleted_files:
        logger.info(f"Detected {len(deleted_files)} deleted document(s)")
        for file_path in deleted_files:
            try:
                vectorstore.delete(where={"source": file_path})
                _update_manifest("delete", file_path)
                logger.info(f"Removed chunks for {os.path.basename(file_path)}")
            except Exception as e:
                logger.error(f"Error deleting {file_path}: {e}")

    files_to_ingest = new_files + updated_files

    if files_to_ingest:
        if new_files:
            logger.info(f"Detected {len(new_files)} new document(s)")
        if updated_files:
            logger.info(f"Detected {len(updated_files)} updated document(s)")

        for file_path in updated_files:
            try:
                vectorstore.delete(where={"source": file_path})
            except Exception:
                pass

        docs = []

        # Optimization: Multiprocessing
        with concurrent.futures.ProcessPoolExecutor() as executor:
            future_to_file = {executor.submit(load_document_by_extension, fp): fp for fp in files_to_ingest}

            for i, future in enumerate(concurrent.futures.as_completed(future_to_file)):
                fp = future_to_file[future]
                try:
                    loaded_docs = future.result()
                    docs.extend(loaded_docs)
                    logger.info(f"[{i + 1}/{len(files_to_ingest)}] Loaded: {os.path.basename(fp)}")
                except Exception as exc:
                    logger.error(f"[{i + 1}/{len(files_to_ingest)}] Failed: {fp}: {exc}")

        if docs:
            logger.info(f"Splitting & embedding {len(docs)} document chunk(s)")
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=1000,
                chunk_overlap=200,
                add_start_index=True
            )
            splits = text_splitter.split_documents(docs)
            splits = _enrich_line_numbers(splits)
            # Stamp tags from manifest onto each split
            for split in splits:
                src = split.metadata.get("source", "")
                t = manifest.get(src, {}).get("tags", [])
                if t:
                    split.metadata["tags"] = ",".join(t)
            splits = filter_complex_metadata(splits)

            vectorstore.add_documents(splits)

            # Update index tracking
            for file_path in files_to_ingest:
                _update_manifest("add", file_path)
            logger.info("Ingestion complete")
    else:
        logger.info("No document changes to ingest")

    # 4. Start Background Sync (Watchdog)
    logger.info("Starting live document watcher")

    # Start Worker Thread
    worker_thread = threading.Thread(target=ingestion_worker, daemon=True)
    worker_thread.start()

    # Start Observer
    event_handler = DocumentEventHandler()
    _OBSERVER = Observer()
    _OBSERVER.schedule(event_handler, DOC_FOLDER, recursive=True)
    _OBSERVER.start()

    return vectorstore


def shutdown() -> None:
    """Stop the ingestion worker and watchdog observer cleanly."""
    INGESTION_QUEUE.put(None)  # sentinel tells the worker to exit
    if _OBSERVER is not None:
        _OBSERVER.stop()
        _OBSERVER.join()
    logger.info("Document engine shutdown complete")


def get_indexed_files() -> dict:
    """Returns the index manifest. Keys are file paths, values are {mtime, size, hash} dicts."""
    return _load_index_manifest()


def get_chunk_counts() -> dict[str, int]:
    """Returns {file_path: chunk_count} by reading ChromaDB metadata in one call."""
    if GLOBAL_VECTORSTORE is None:
        return {}
    try:
        results = GLOBAL_VECTORSTORE.get(include=["metadatas"])
        counts: dict[str, int] = {}
        for meta in results.get("metadatas", []):
            source = meta.get("source", "")
            if source:
                counts[source] = counts.get(source, 0) + 1
        return counts
    except Exception as e:
        logger.warning(f"Could not fetch chunk counts: {e}")
        return {}


def trigger_reindex() -> int:
    """
    Queues every file in the input folder for re-ingestion via the background worker.
    Returns the number of files queued.
    """
    count = 0
    for root, _, files in os.walk(DOC_FOLDER):
        for file in files:
            if file.startswith("~$"):
                continue
            if file.lower().endswith(SUPPORTED_EXTENSIONS):
                INGESTION_QUEUE.put(("update", os.path.join(root, file)))
                count += 1
    return count


def get_retriever(k=4, source_filter: Optional[str] = None, tag_filter: Optional[str] = None):
    """Returns a retriever interface from the current global vectorstore.

    Args:
        k: Number of documents to retrieve.
        source_filter: If provided, restricts results to documents whose
            ``source`` metadata field matches this full file path.
        tag_filter: If provided, restricts results to documents whose
            ``tags`` metadata field contains this tag string.
    """
    global GLOBAL_VECTORSTORE
    if GLOBAL_VECTORSTORE is None:
        initialize_vectorstore()

    search_kwargs: dict = {"k": k, "fetch_k": 20, "lambda_mult": 0.5}
    if source_filter:
        search_kwargs["filter"] = {"source": source_filter}
    if tag_filter:
        tag_f = {"tags": {"$contains": tag_filter}}
        if "filter" in search_kwargs:
            search_kwargs["filter"] = {"$and": [search_kwargs["filter"], tag_f]}
        else:
            search_kwargs["filter"] = tag_f

    return GLOBAL_VECTORSTORE.as_retriever(
        search_type="mmr",
        search_kwargs=search_kwargs,
    )


# --- PART 2: MODELS & PROMPTS ---

# Initialize LLMs
thinking_llm = ChatOllama(model=THINKING_OLLAMA_MODEL, temperature=0)
fast_llm = ChatOllama(model=FAST_OLLAMA_MODEL, temperature=0)


# Grader
class GradeDocuments(BaseModel):
    """Binary score for relevance check on retrieved documents."""
    binary_score: str = Field(description="Documents are relevant to the question, 'yes' or 'no'")


structured_llm_grader = fast_llm.with_structured_output(GradeDocuments)
grader_system = """You are a grader assessing whether a retrieved document contains information needed to answer a user question.
Grade as 'yes' ONLY if the document directly contains information that addresses the specific question — not merely shares the same general topic or setting.
Grade as 'no' if the document's content is tangential or does not actually help answer the question.
Give a binary score 'yes' or 'no'."""
grader_prompt = ChatPromptTemplate.from_messages(
    [("system", grader_system), ("human", "Retrieved document: \n\n {document} \n\n User question: {question}")]
)
grader_chain = grader_prompt | structured_llm_grader

rag_system_prompt = """You are a helpful assistant. Answer the user's question based ONLY on the context provided below.
Each context chunk is labelled with a reference number: [1], [2], etc.

Rules:
1. Base your answer SOLELY on what is explicitly written in the provided chunks.
2. Cite a chunk [N] inline ONLY if the specific fact you are stating appears literally in that chunk's text.
3. Do NOT attribute information to a source if that information does not appear in that source's text.
4. If the context does not contain the answer, say "I don't know."
5. Do NOT add a Sources list at the end — one will be appended automatically.
"""

# RAG Generator
rag_prompt = ChatPromptTemplate.from_messages(
    [
        ("system", rag_system_prompt),
        ("human", "Question: {question} \n\n Context: {context} \n\n Answer:"),
    ]
)
rag_chain = rag_prompt | thinking_llm | StrOutputParser()

# Rewriter
rewrite_system = "You are a question re-writer that converts an input question to a better version for vector retrieval."
rewrite_prompt = ChatPromptTemplate.from_messages(
    [("system", rewrite_system), ("human", "Initial question: \n\n {question} \n Formulate an improved question.")]
)
rewriter_chain = rewrite_prompt | fast_llm | StrOutputParser()


# --- PART 3: GRAPH NODES (OPTIMIZED) ---

class GraphState(TypedDict):
    question: str
    generation: str
    documents: List[Document]
    loop_step: int
    scoped_source: NotRequired[Optional[str]]         # None = search all documents
    stream_queue: NotRequired[Optional[queue.Queue]]  # queue for live token streaming
    tag_filter: NotRequired[Optional[str]]            # optional tag to restrict retrieval


def retrieve(state):
    question = state["question"]
    scoped_source = state.get("scoped_source")
    tag_filter = state.get("tag_filter")
    logger.debug(f"[RETRIEVE] question: {question!r}" +
                 (f" | scoped to: {os.path.basename(scoped_source)}" if scoped_source else ""))

    retriever = get_retriever(k=4, source_filter=scoped_source, tag_filter=tag_filter)
    documents = retriever.invoke(question)

    sources = [os.path.basename(d.metadata.get("source", "?")) for d in documents]
    logger.debug(f"[RETRIEVE] got {len(documents)} doc(s): {sources}")
    return {"documents": documents, "question": question}


def grade_documents(state):
    question = state["question"]
    documents = state["documents"]
    logger.debug(f"[GRADE] grading {len(documents)} doc(s) against question: {question!r}")

    if not documents:
        logger.debug("[GRADE] no documents to grade")
        return {"documents": [], "question": question}

    batch_inputs = [{"question": question, "document": d.page_content} for d in documents]
    scores = grader_chain.batch(batch_inputs)

    filtered_docs = []
    for i, score in enumerate(scores):
        source = os.path.basename(documents[i].metadata.get("source", "?"))
        verdict = score.binary_score
        logger.debug(f"[GRADE] doc {i + 1} ({source}): {verdict.upper()}")
        if verdict == "yes":
            filtered_docs.append(documents[i])

    logger.debug(f"[GRADE] {len(filtered_docs)}/{len(documents)} docs passed")
    return {"documents": filtered_docs, "question": question}


def generate_rag(state):
    question = state["question"]
    documents = state["documents"]

    formatted_context_list = []
    for i, doc in enumerate(documents):
        source_name = os.path.basename(doc.metadata.get("source", "Unknown"))
        location = _citation_location(doc.metadata)
        loc_str = f" ({location})" if location else ""
        formatted_context_list.append(f"[{i + 1}] {source_name}{loc_str}\n{doc.page_content}")

    full_context_string = "\n\n---\n\n".join(formatted_context_list)

    sq: Optional[queue.Queue] = state.get("stream_queue")

    logger.debug(
        f"[GENERATE] → {THINKING_OLLAMA_MODEL}"
        f"\n  question : {question!r}"
        f"\n  context  : {len(full_context_string)} chars across {len(documents)} chunk(s)"
        f"\n{full_context_string}"
    )

    generation = ""
    for chunk in rag_chain.stream({"context": full_context_string, "question": question}):
        generation += chunk
        if sq is not None:
            sq.put(chunk)

    # Append a Sources section using page → line priority
    if documents:
        source_lines = []
        for i, doc in enumerate(documents):
            meta = doc.metadata
            fname = os.path.basename(meta.get("source", "Unknown"))
            loc = _citation_location(meta)
            source_lines.append(f"[{i + 1}] {fname}" + (f" — {loc}" if loc else ""))
        footer = "\n\n---\n**Sources:**\n" + "\n".join(source_lines)
        generation += footer
        if sq is not None:
            sq.put(footer)

    if sq is not None:
        sq.put(None)  # sentinel — signals streaming is complete

    logger.debug(f"[GENERATE] ← response:\n{generation}")
    return {"generation": generation, "documents": documents}


def transform_query(state):
    question = state["question"]
    documents = state["documents"]
    loop_step = state.get("loop_step", 0)
    logger.debug(f"[REWRITE] → {FAST_OLLAMA_MODEL} | attempt {loop_step + 1} | input: {question!r}")
    better_question = rewriter_chain.invoke({"question": question})
    logger.debug(f"[REWRITE] ← output: {better_question!r}")
    return {"documents": documents, "question": better_question, "loop_step": loop_step + 1}


def decide_to_generate(state):
    filtered_documents = state["documents"]
    loop_step = state.get("loop_step", 0)

    if not filtered_documents:
        if loop_step >= 3:
            return "generate"
        return "transform_query"
    return "generate"


# --- PART 4: WORKFLOW BUILD ---

workflow = StateGraph(GraphState)
workflow.add_node("retrieve", retrieve)
workflow.add_node("grade_documents", grade_documents)
workflow.add_node("generate_rag", generate_rag)
workflow.add_node("transform_query", transform_query)

workflow.add_edge(START, "retrieve")
workflow.add_edge("retrieve", "grade_documents")
workflow.add_conditional_edges(
    "grade_documents",
    decide_to_generate,
    {"transform_query": "transform_query", "generate": "generate_rag"},
)
workflow.add_edge("transform_query", "retrieve")
workflow.add_edge("generate_rag", END)

app = workflow.compile()


# --- ENTRY POINTS ---

def _run_rag_graph(inputs: dict, include_sources: bool):
    """Execute the compiled RAG graph and return a formatted answer."""
    config = {"recursion_limit": 25}
    final_state = None
    generate_rag_state = None

    try:
        for output in app.stream(inputs, config=config):
            for key, value in output.items():
                final_state = value
                if key == "generate_rag":
                    generate_rag_state = value
    except Exception as e:
        logger.error(f"Graph execution error: {e}")
        return f"Error: {e}" if not include_sources else {"error": str(e)}

    result_state = generate_rag_state if generate_rag_state is not None else final_state
    if result_state is None:
        no_answer = "No answer generated."
        return no_answer if not include_sources else {"answer": no_answer, "citations": []}

    final_answer = result_state.get("generation", "No answer generated.")
    used_docs = result_state.get("documents", [])

    sources_data = []
    for doc in used_docs:
        meta = doc.metadata
        fname = os.path.basename(meta.get("source", "Unknown"))
        location = _format_location(meta)
        sources_data.append({
            "file": fname,
            "location": location,
            "snippet": doc.page_content[:50] + "...",
        })

    if include_sources:
        return {"answer": final_answer, "citations": sources_data}
    return final_answer


def query_documents(user_input: str, include_sources: bool = False,
                    stream_queue: Optional[queue.Queue] = None,
                    tag_filter: Optional[str] = None):
    """Search all indexed documents and return an answer."""
    if GLOBAL_VECTORSTORE is None:
        initialize_vectorstore()
    inputs: dict = {"question": user_input, "loop_step": 0}
    if stream_queue is not None:
        inputs["stream_queue"] = stream_queue
    if tag_filter:
        inputs["tag_filter"] = tag_filter
    return _run_rag_graph(inputs, include_sources)


def query_documents_scoped(user_input: str, filename: str, include_sources: bool = False,
                           stream_queue: Optional[queue.Queue] = None,
                           tag_filter: Optional[str] = None):
    """Search within a single named document and return an answer.

    Args:
        user_input: The user's question.
        filename: The basename of the indexed file to restrict the search to.
        include_sources: If True, return a dict with 'answer' and 'citations'.
        stream_queue: Optional queue that receives token chunks as they are generated.
        tag_filter: Optional tag string to further restrict the search.
    """
    if GLOBAL_VECTORSTORE is None:
        initialize_vectorstore()

    manifest = _load_index_manifest()
    matching = [p for p in manifest if os.path.basename(p) == filename]
    if not matching:
        msg = f"No indexed document named '{filename}' found. Use `/status` to see available documents."
        return {"answer": msg, "citations": []} if include_sources else msg

    inputs: dict = {"question": user_input, "loop_step": 0, "scoped_source": matching[0]}
    if stream_queue is not None:
        inputs["stream_queue"] = stream_queue
    if tag_filter:
        inputs["tag_filter"] = tag_filter
    return _run_rag_graph(inputs, include_sources)


def similarity_search(
    query: str,
    k: int = 10,
    source_filter: Optional[str] = None,
    tag_filter: Optional[str] = None,
) -> list[tuple]:
    """Return top-k (Document, relevance_score) pairs without LLM generation.

    relevance_score is in [0, 1] — higher means more similar.
    source_filter, if given, must be the full file path stored in the manifest.
    Returns an empty list if the vectorstore is not ready or an error occurs.
    """
    if GLOBAL_VECTORSTORE is None:
        return []
    filter_dict = {"source": source_filter} if source_filter else None
    if tag_filter:
        tag_f = {"tags": {"$contains": tag_filter}}
        if filter_dict is not None:
            filter_dict = {"$and": [filter_dict, tag_f]}
        else:
            filter_dict = tag_f
    try:
        return GLOBAL_VECTORSTORE.similarity_search_with_relevance_scores(
            query, k=k, filter=filter_dict
        )
    except Exception as e:
        logger.error(f"Similarity search error: {e}")
        return []


def get_chunks_for_file(file_path: str) -> list[dict]:
    """Return all stored chunks for a file as a list of {content, metadata} dicts.

    Results are sorted by start_index when available, otherwise by page.
    Returns an empty list if the vectorstore is not ready or an error occurs.
    """
    if GLOBAL_VECTORSTORE is None:
        return []
    try:
        results = GLOBAL_VECTORSTORE.get(
            where={"source": file_path},
            include=["metadatas", "documents"],
        )
        chunks = [
            {"content": doc, "metadata": meta}
            for doc, meta in zip(
                results.get("documents", []),
                results.get("metadatas", []),
            )
        ]
        chunks.sort(key=lambda c: (
            c["metadata"].get("page", 0),
            c["metadata"].get("start_index", 0),
        ))
        return chunks
    except Exception as e:
        logger.error(f"Error fetching chunks for {file_path}: {e}")
        return []


if __name__ == "__main__":
    query_documents("Who is best friends with Dagbert?", True)
